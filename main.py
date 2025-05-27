from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook 
from email.message import EmailMessage
import smtplib 


def initialize_driver():
    service = Service()
    options = webdriver.ChromeOptions()
    return webdriver.Chrome(service=service, options=options)


def scrape_data(driver, url, max_price, number_of_rooms):
    driver.get(url)

    wait = WebDriverWait(driver, 10)

    wait.until(EC.presence_of_element_located((By.ID, "mtd_59"))).click()
    wait.until(EC.presence_of_element_located((By.ID, "ahc_14195"))).click()
    wait.until(EC.presence_of_element_located((By.ID, "ahc_1088"))).click()
    find = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "l100")))
    Select(find).select_by_visible_text("Pārdod")
    find = wait.until(EC.presence_of_element_located((By.NAME, "topt[1][min]")))
    Select(find).select_by_visible_text(str(number_of_rooms))
    find = wait.until(EC.presence_of_element_located((By.NAME, "topt[1][max]")))
    Select(find).select_by_visible_text(str(number_of_rooms))
    wait.until(EC.presence_of_element_located((By.ID, "f_o_8_max"))).send_keys(max_price)
    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "s12"))).click()

    table = wait.until(EC.presence_of_element_located((By.XPATH, "//form/table[@align='center']")))
    rows = table.find_elements(By.TAG_NAME, "tr")[1:-1]

    table_data = []

    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")[2:]

        link = cells[0].find_element(By.TAG_NAME, "a")
        row_data = [link.get_attribute('href')] + [cell.text for cell in cells]

        table_data.append(row_data)

    return table_data


def save_to_excel(data, file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            ws.cell(row=i, column=j, value=value)

    wb.save(file_path)
    wb.close()


def send_email(file_path, my_email, my_password, recipient_email):
    smtp_server = "mail.inbox.lv"
    smtp_port = 587

    try:
        my_server = smtplib.SMTP(smtp_server, smtp_port)
        my_server.ehlo()
        my_server.starttls()

        my_server.login(my_email, my_password)

        msg = EmailMessage()
        msg.set_content('Apartments in Agenskalns:')

        msg['Subject'] = 'Apartments'
        msg['From'] = my_email
        msg['To'] = recipient_email

        with open(file_path, 'rb') as file:
            file_data = file.read()
            msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file.name)

        my_server.send_message(msg)
        print('Mail Sent')

    except Exception as e:
        print(f'Error: {e}')

    finally:
        my_server.quit()


url = "https://ss.lv"
file_path = "apartments-list.xlsx"
max_price = 80000
number_of_rooms = 2
my_email = "dip225project@inbox.lv"
my_password = ""
recipient_email = "dip225project@inbox.lv"

driver = initialize_driver()

try:
    table_data = scrape_data(driver, url, max_price, number_of_rooms)

    save_to_excel(table_data, file_path)
    send_email(file_path, my_email, my_password, recipient_email)

    print(table_data)

finally:
    driver.quit()
